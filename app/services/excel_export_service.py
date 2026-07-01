"""
Excel Export Service — generates a multi-sheet XLSX workbook from harvest results.

Workbook structure
──────────────────
Sheet 1 — Combined Jobs   (all sources, all fields + lead intelligence columns)
Sheet 2 — LinkedIn Jobs
Sheet 3 — Naukri Jobs
Sheet 4 — Dice Jobs
Sheet 5 — Lead Intelligence   (all jobs with lead status column)

Output path:  data/results/excel/YYYYMMDD_HHMMSS_harvest.xlsx
"""
from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import structlog

# openpyxl rejects XML control characters — strip them from every cell value
_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f￾￿]')

logger = structlog.get_logger(__name__)

_EXCEL_DIR = Path("data/results/excel")

# ── Column definitions ────────────────────────────────────────────────────────

_JOB_COLUMNS: list[tuple[str, str]] = [
    ("job_title",              "Job Title"),
    ("company",                "Company"),
    ("location",               "Location"),
    ("salary",                 "Salary"),
    ("experience",             "Experience"),
    ("posted_date",            "Posted Date"),
    ("job_url",                "Job URL"),
    ("job_description",        "Job Description"),
    ("skills",                 "Skills"),
    ("work_mode",              "Work Mode"),
    ("source",                 "Source"),
    ("job_type",               "Job Type"),
    ("domain",                 "Domain"),
    ("hiring_entity",          "Hiring Entity"),
    ("is_gcc",                 "Is GCC"),
    ("verification_status",    "Verification Status"),
    ("job_poster_name",        "Job Poster Name"),
    ("job_poster_designation", "Job Poster Designation"),
    ("linkedin_profile_url",   "LinkedIn Profile URL"),
    ("current_company",        "Current Company"),
    ("email_id",               "Email ID"),
    ("contact_number",         "Contact Number"),
]

_LEAD_COLUMNS: list[tuple[str, str]] = [
    ("job_title",              "Job Title"),
    ("company",                "Company"),
    ("source",                 "Source"),
    ("job_poster_name",        "Job Poster Name"),
    ("job_poster_designation", "Job Poster Designation"),
    ("linkedin_profile_url",   "LinkedIn Profile URL"),
    ("current_company",        "Current Company"),
    ("email_id",               "Email ID"),
    ("contact_number",         "Contact Number"),
    ("hiring_entity",          "Hiring Entity"),
    ("verification_status",    "Verification Status"),
    ("job_url",                "Job URL"),
    ("posted_date",            "Posted Date"),
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sanitize(val: Any) -> Any:
    """Strip illegal XML control characters so openpyxl never raises IllegalCharacterError."""
    if isinstance(val, str):
        val = _ILLEGAL_XML.sub("", val)
        # Truncate very long strings (job descriptions) to 5000 chars
        if len(val) > 5000:
            val = val[:5000] + "…"
    return val


def _lead_status(job: Any) -> str:
    """Compute Lead Status for a job record."""
    has_name  = bool(getattr(job, "job_poster_name", None))
    has_email = bool(getattr(job, "email_id", None))
    has_phone = bool(getattr(job, "contact_number", None))
    has_url   = bool(getattr(job, "linkedin_profile_url", None))
    if has_email or has_phone:
        return "Enriched - Contact Available"
    if has_name and (has_url or getattr(job, "current_company", None)):
        return "Enriched - Profile Only"
    if has_name:
        return "Partial - Name Only"
    return "Pending"


def _job_to_row(job: Any, columns: list[tuple[str, str]]) -> list[Any]:
    """Convert a UnifiedJob to a flat list aligned with `columns`."""
    row: list[Any] = []
    for field_key, _ in columns:
        if field_key == "_lead_status":
            val = _lead_status(job)
        else:
            val = getattr(job, field_key, None)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif val is None:
                val = ""
        row.append(_sanitize(val))
    return row


def _apply_header_style(ws: Any, header_fill: str = "1F4E79") -> None:
    """Bold white text on dark-blue background for the header row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill("solid", fgColor=header_fill)
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(wrap_text=True, vertical="center")
        for cell in ws[1]:
            cell.fill  = fill
            cell.font  = font
            cell.alignment = align
    except Exception:
        pass  # openpyxl not available — skip styling


def _autofit_columns(ws: Any, max_width: int = 60) -> None:
    """Set reasonable column widths based on content."""
    try:
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value or "")) for cell in col_cells),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
    except Exception:
        pass


def _write_sheet(ws: Any, columns: list[tuple[str, str]], jobs: list[Any]) -> None:
    """Write header + data rows to a worksheet."""
    ws.append([display for _, display in columns])
    for job in jobs:
        ws.append(_job_to_row(job, columns))
        logger.debug("excel_row_written", sheet=ws.title, rows=ws.max_row - 1)
    _apply_header_style(ws)
    _autofit_columns(ws)
    ws.freeze_panes = "A2"
    logger.info("sheet_completed", sheet=ws.title, rows=ws.max_row - 1)


# ══════════════════════════════════════════════════════════════════════════════
# ExcelExportService
# ══════════════════════════════════════════════════════════════════════════════

class ExcelExportService:
    """Generate a multi-sheet Excel workbook from harvest results."""

    def export(
        self,
        all_jobs:       list[Any],
        jobs_by_source: dict[str, list[Any]],
        run_id:         str,
        filters_snap:   dict,
    ) -> str:
        """
        Build the workbook and write it to data/results/excel/.
        Returns the absolute path of the saved file.
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("excel_export_skipped", reason="openpyxl not installed — run: pip install openpyxl")
            return ""

        _EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        ts       = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"{ts}_harvest.xlsx"
        path     = _EXCEL_DIR / filename

        wb = openpyxl.Workbook()

        # ── Sheet 1: Combined Jobs ─────────────────────────────────────────────
        ws_combined = wb.active
        ws_combined.title = "Combined Jobs"
        _write_sheet(ws_combined, _JOB_COLUMNS, all_jobs)

        # ── Sheets 2-4: per-source ─────────────────────────────────────────────
        for sheet_title in ("LinkedIn Jobs", "Naukri Jobs", "Dice Jobs"):
            source_key = sheet_title.replace(" Jobs", "")  # "LinkedIn" | "Naukri" | "Dice"
            ws = wb.create_sheet(title=sheet_title)
            source_jobs = jobs_by_source.get(source_key, [])
            if source_key == "LinkedIn":
                logger.info("linkedin_sheet_created", sheet=sheet_title, job_count=len(source_jobs))
            _write_sheet(ws, _JOB_COLUMNS, source_jobs)
            if source_key == "LinkedIn":
                logger.info("linkedin_jobs_written_to_excel", rows=len(source_jobs))

        # ── Sheet 5: Lead Intelligence (ALL jobs with lead status column) ────
        ws_leads = wb.create_sheet(title="Lead Intelligence")
        _write_sheet(ws_leads, _LEAD_COLUMNS, all_jobs)

        enriched_count = sum(
            1 for j in all_jobs
            if getattr(j, "job_poster_name", None)
            or getattr(j, "email_id", None)
            or getattr(j, "contact_number", None)
        )

        wb.save(str(path))
        logger.info(
            "excel_exported",
            path         = str(path.resolve()),
            total_jobs   = len(all_jobs),
            lead_records = enriched_count,
            sheets       = ["Combined Jobs", "LinkedIn Jobs", "Naukri Jobs", "Dice Jobs", "Lead Intelligence"],
        )
        logger.info("harvest_completed", total_jobs=len(all_jobs), lead_records=enriched_count, excel_generated=True)
        return str(path.resolve())

    def export_path_for_run(self, run_id: str) -> str:
        """Return the expected Excel path for a run_id (may not exist yet)."""
        _EXCEL_DIR.mkdir(parents=True, exist_ok=True)
        return str((_EXCEL_DIR / f"{run_id}_harvest.xlsx").resolve())
