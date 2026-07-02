"""
Lead Excel Service — exports LeadIntelligenceResult to a styled .xlsx file.

Column layout (19 columns)
───────────────────────────
A  Lead ID            J  Official Email
B  Recruiter Name     K  Email Status
C  Designation        L  Contact Number
D  Department         M  Phone Status
E  Company            N  Employment History (joined)
F  Current Company    O  Source
G  Location           P  Confidence Score
H  LinkedIn URL       Q  Last Verified
I  Job Post URL       R  CRM Status
                      S  Confidence Value (internal — hidden col)
"""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from app.models.lead_models import LeadIntelligenceResult

import structlog

logger = structlog.get_logger(__name__)

_OUTPUT_DIR = Path("data/results/lead_intelligence")

# Column definitions: (header, field_accessor)
_COLUMNS: list[tuple[str, str]] = [
    ("Lead ID",            "lead_id"),
    ("Recruiter Name",     "recruiter_name"),
    ("Designation",        "designation"),
    ("Department",         "department"),
    ("Company",            "company"),
    ("Current Company",    "current_company"),
    ("Location",           "location"),
    ("LinkedIn URL",       "linkedin_profile_url"),
    ("Job Post URL",       "job_post_url"),
    ("Official Email",     "official_email"),
    ("Email Status",       "email_status"),
    ("Contact Number",     "contact_number"),
    ("Phone Status",       "phone_status"),
    ("Employment History", "_employment_history_str"),
    ("Source",             "_source_str"),
    ("Confidence Score",   "confidence_score"),
    ("Last Verified",      "last_verified"),
    ("CRM Status",         "crm_status"),
]


class LeadExcelService:
    """Exports a LeadIntelligenceResult to a formatted .xlsx workbook."""

    def export(self, result: "LeadIntelligenceResult", run_id: str) -> str:
        """
        Write leads to Excel and return the resolved file path.

        Requires openpyxl; raises ImportError if not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment,
                Font,
                PatternFill,
            )
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("openpyxl is required for Excel export") from exc

        _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        ts   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        path = _OUTPUT_DIR / f"{run_id}_{ts}_leads.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lead Intelligence"

        # ── Summary sheet header block ─────────────────────────────────────────
        self._write_summary_block(ws, result, Font, PatternFill, Alignment)

        header_row = 8  # summary takes rows 1-7; headers start at row 8

        # ── Column headers ────────────────────────────────────────────────────
        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="1F4E79")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (header, _) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align

        ws.row_dimensions[header_row].height = 30

        # ── Data rows ─────────────────────────────────────────────────────────
        fill_high   = PatternFill("solid", fgColor="E2EFDA")  # light green
        fill_medium = PatternFill("solid", fgColor="FFEB9C")  # light yellow
        fill_low    = PatternFill("solid", fgColor="FFC7CE")  # light red
        normal_align = Alignment(vertical="top", wrap_text=True)

        confidence_fill = {
            "High":   fill_high,
            "Medium": fill_medium,
            "Low":    fill_low,
        }

        for row_offset, lead in enumerate(result.leads, start=1):
            row = header_row + row_offset
            row_fill = confidence_fill.get(lead.confidence_score, PatternFill())

            for col_idx, (_, field) in enumerate(_COLUMNS, start=1):
                if field == "_employment_history_str":
                    value = " | ".join(lead.employment_history) if lead.employment_history else ""
                elif field == "_source_str":
                    value = ", ".join(lead.source) if lead.source else ""
                else:
                    value = getattr(lead, field, "")

                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.alignment = normal_align

        # ── Column widths ─────────────────────────────────────────────────────
        widths = [18, 22, 25, 22, 22, 22, 18, 40, 40, 30, 14, 18, 14, 45, 22, 16, 22, 12]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── Freeze panes at data start ────────────────────────────────────────
        ws.freeze_panes = f"A{header_row + 1}"

        # ── Auto-filter ───────────────────────────────────────────────────────
        last_col = get_column_letter(len(_COLUMNS))
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(result.leads)}"

        wb.save(path)
        logger.info("lead_excel_written", path=str(path), rows=len(result.leads))
        return str(path.resolve())

    # ── Summary block ──────────────────────────────────────────────────────────

    def _write_summary_block(
        self, ws: object, result: "LeadIntelligenceResult",
        Font: object, PatternFill: object, Alignment: object,
    ) -> None:
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align

        title_font   = _Font(bold=True, size=13, color="FFFFFF")
        title_fill   = _Fill("solid", fgColor="1F4E79")
        label_font   = _Font(bold=True, size=10)
        value_font   = _Font(size=10)
        center_align = _Align(horizontal="center", vertical="center")

        ws.merge_cells("A1:R1")
        title_cell = ws["A1"]
        title_cell.value     = "Hybrid Lead Intelligence Report"
        title_cell.font      = title_font
        title_cell.fill      = title_fill
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 28

        summary_rows = [
            ("Run ID",                 result.run_id),
            ("Keyword",                result.keyword),
            ("Executed At",            result.executed_at),
            ("Total Leads",            result.total_leads),
            ("High / Medium / Low",    f"{result.high_confidence} / {result.medium_confidence} / {result.low_confidence}"),
            ("LinkedIn Posts Found",   result.linkedin_posts_found),
            ("Naukri Fallbacks",       result.premium_naukri_fallbacks),
        ]

        for row_i, (label, value) in enumerate(summary_rows, start=2):
            lc = ws.cell(row=row_i, column=1, value=label)
            lc.font      = label_font
            lc.alignment = _Align(horizontal="right", vertical="center")

            vc = ws.cell(row=row_i, column=2, value=str(value))
            vc.font      = value_font
            vc.alignment = _Align(horizontal="left", vertical="center")

            ws.merge_cells(f"B{row_i}:R{row_i}")
            ws.row_dimensions[row_i].height = 16
