"""
One-time script: create data/prospects/input/prospects.xlsx
from the R&D prospect list.

Run once:  python create_prospects_xlsx.py
"""
import openpyxl
from pathlib import Path

PROSPECTS = [
    ("Kiya AI",                                        "Rajesh Mirjankar",          "MD & CEO"),
    ("Kiya AI",                                        "Prem Kumar",                "Sr. HR"),
    ("Fino Payments Bank",                             "Pratima Pinto Thomas",      "HR Head"),
    ("Fino Payments Bank",                             "Ravish Rathore",            "Lead Human Resources Business Partner"),
    ("Fino Payments Bank",                             "Nicholas Crasto",           "Deputy VP - TA"),
    ("Odessa",                                         "Rashmi Jagannath",          "VP HR"),
    ("Odessa",                                         "Shilpa Jadav",              "Head People Success Team"),
    ("Amadeus Labs",                                   "Aanchal Barua",             "Cluster Lead Talent Acquisition, Singapore & North Asia"),
    ("Amadeus Labs",                                   "Leetha Prajesh",            "Senior Director-Talent Acquisition, APAC and India"),
    ("Amadeus Labs",                                   "Shankar Padhy",             "TA Lead"),
    ("Amadeus Labs",                                   "Ruchika Kumari",            "Manager - Recruitment"),
    ("Ness Digital Engg",                              "Vijendra Vaishnav",         "Sr. Manager TA"),
    ("Ness Digital Engg",                              "Vipasha Soni",              "TA Leader"),
    ("Ness Digital Engg",                              "Ravish Chadha",             "Sr. Director - TA"),
    ("Ness Digital Engg",                              "Toral Sanghavi",            "Associate VP & Head of HR"),
    ("Fractal Analytics",                              "Veena Brahmakal",           "VP | Global Head of Talent Management-Staffing"),
    ("Fractal Analytics",                              "Savita Hortikar",           "VP | Global Head of Talent Management"),
    ("Fractal Analytics",                              "Sony Alex",                 "Manager TA"),
    ("Fractal Analytics",                              "Manjunath Hedge",           "Senior Manager - Talent Acquisition"),
    ("Fractal Analytics",                              "Sakthi Kumar",              "Senior Manager | AI & Leadership Hiring"),
    ("Lentra AI",                                      "Ashish A Potdar",           "VP HR"),
    ("Perfios",                                        "Darshan Bora",              "VP - HR"),
    ("Perfios",                                        "Pratima Kumari",            "Talent Scout - People Success Team"),
    ("Perfios",                                        "Shreya Tyagi",              "TA Specialist"),
    ("Perfios",                                        "Joyce Felicia",             "Sr. Director HR"),
    ("Perfios",                                        "Menu Sharma",               "Sr. People Success Partner"),
    ("Perfios",                                        "Tushna Singanporia",        "AVP - Talent Management"),
    ("Perfios",                                        "A Vinith",                  "Manager HR"),
    ("Oracle Financial Services Software Limited",     "Kavita Pinto",              "Sr HR Manager"),
    ("Mastek",                                         "Sheena Persis Fedrick James","Senior Talent Acquisition Specialist"),
    ("Mastek",                                         "Kapil Katira",              "Sr. Manager TA"),
    ("Mastek",                                         "Divyesh Makwana",           "Talent Acquisition Manager"),
    ("Mastek",                                         "Rajat Adhikari",            "VP Global Head - TA"),
    ("Mastek",                                         "Purva Mudgal",              "Global Talent Management Manager"),
    ("Mastek",                                         "Jane Maria James",          "HR Consultant"),
    ("Mastek",                                         "Priyanka Gaur",             "TA Specialist"),
    ("Mastek",                                         "Tushar Bhave",              "Sr. Manager - TA"),
    ("Mastek",                                         "Sneha Kumari",              "Manager - TA"),
    ("Mastek",                                         "Sathya Muthuchellam",       "TA Manager"),
    ("Mastek",                                         "Sweta Shukla",              "Recruitment Manager"),
    ("Mastek",                                         "Ruchi Prakash Joshi",       "Associate Manager"),
    ("Ciklum",                                         "Balaji Prabhakaran",        "Associate Director - TA"),
    ("Ciklum",                                         "Monika Rathan",             "Manager - TA"),
    ("Ciklum",                                         "Samrudhi Pardhe",           "Sr. TA Specialist"),
    ("Ciklum",                                         "Jerin Joshua",              "Recruitment Consultant"),
    ("Ciklum",                                         "Persis (Amita) A",          "Sr TA"),
    ("Ciklum",                                         "Hari Baskar",               "Recruitment Consultant"),
    ("Ciklum",                                         "Gopikrishnan P",            "Sr Recruiter"),
    ("Ciklum",                                         "Akansha Raipatrewar",       "HR Executive"),
    ("Ciklum",                                         "Nandhini N",                "Sr. TA Specialist"),
    ("Ciklum",                                         "Jothilingam K",             "Associate Manager - TA"),
    ("Ciklum",                                         "Sourabh Baranj",            "Recruitment Consultant"),
    ("Hexaware",                                       "Sangeetha G",               "VP - Head of TA"),
    ("Hexaware",                                       "Ashish Dua",                "Sr. GM - TA"),
    ("Hexaware",                                       "Raja Pandiyan",             "Manager - TA"),
    ("Hexaware",                                       "Manjunath GM",              "TA Leader"),
    ("Hexaware",                                       "Jitendra Bhatauria",        "TA"),
    ("Hexaware",                                       "Mahalakshmi",               "Sr. Executive - TA"),
    ("Hexaware",                                       "Kalpana S",                 "TA Specialist"),
    ("E2 Open",                                        "Edward Fransis",            "Head of TA"),
    ("E2 Open",                                        "Afsana Shaik",              "TA Specialist"),
    ("E2 Open",                                        "Srimala Sarkar Chandra",    "HRBP"),
    ("E2 Open",                                        "Pratibha Kakar",            "HR Head"),
    ("DXC Technologies",                               "Ramya MC",                  "Talent Fulfillment Strategist"),
    ("DXC Technologies",                               "Bhavika Deshmukh",          "TA Specialist"),
    ("DXC Technologies",                               "Jaya Raghuvanshi",          "Recruitment Relationship Manager"),
    ("DXC Technologies",                               "Shivani Barupal",           "Sr. Relationship Recruitment Manager"),
    ("DXC Technologies",                               "Pratibha Sharma",           "TA"),
    ("DXC Technologies",                               "Vaibhav Mahajan",           "TA Head"),
    ("Niyoto Infotech",                                "Mamta Chauhan",             "Sr. TA Specialist"),
    ("Niyoto Infotech",                                "Sachin Sharma",             "Sr. TA Specialist"),
    ("Niyoto Infotech",                                "Pavan Kumar R",             "GM - HR"),
    ("Niyoto Infotech",                                "Jenson Mathew",             "Sr Manager TA"),
    ("Niyoto Infotech",                                "Vidya Bhaskaran",           "Manager Recruitment"),
    ("Niyoto Infotech",                                "Fahim Ahmed",               "Sr. TA Specialist"),
    ("Niyoto Infotech",                                "Anupama Sharma",            "Sr. TA Specialist"),
    ("Niyoto Infotech",                                "Gunashekaran",              "Vendor Manager"),
    ("Niyoto Infotech",                                "Mahima Sharma",             "Principal Consultant"),
    ("Niyoto Infotech",                                "Urmila Khobragade",         "Sr. Recruitment Executive"),
    ("HashRoot Limited",                               "Maheswari M",               "TA Specialist"),
    ("HashRoot Limited",                               "Jismi Jiju",                "TA Specialist"),
    ("HashRoot Limited",                               "Shishira AC",               "Talent Acquisition Specialist"),
    ("HashRoot Limited",                               "Afsana Salam",              "Talent Acquisition Specialist"),
    ("HashRoot Limited",                               "Chandana K",                "Talent Acquisition Manager"),
    ("Collabera",                                      "Mihir Seth",                "Associate VP"),
    ("Collabera",                                      "Kishen Patel",              "Sr. Talent Acquisition Specialist"),
    ("Collabera",                                      "Ankur Singh",               "Sr. Talent Acquisition Specialist"),
    ("Collabera",                                      "Vigesh Vyas",               "Sr. Talent Acquisition Specialist"),
    ("Collabera",                                      "Jayendra Desai",            "Sr. Manager TA"),
    ("Collabera",                                      "Jigyasa Shah",              "Sr. Talent Acquisition Specialist"),
    ("Collabera",                                      "Mohammed Haq",              "Account Manager"),
    ("Collabera",                                      "Neel Patel",                "Talent Specialist"),
    ("Collabera",                                      "Vashnavi Reddy",            "Talent Specialist"),
    ("Collabera",                                      "Jignesh Chawda",            "Associate Manager - Vendor Management"),
    ("Collabera",                                      "Moulin Desai",              "Associate VP TA"),
    ("M360 Research",                                  "Sagaya Raj Austin",         "Manager TA"),
    ("Vlink Inc",                                      "Rohit Chaurasia",           "TA Lead"),
    ("Paytm",                                          "Reema Shibu",               "TA"),
    ("Ema Unlimited",                                  "Divya Vasta",               "TA Lead"),
    ("Awign",                                          "Manas Mahapatra",           "Manager TA"),
    ("Hurix",                                          "Paramita Sengupta",         "Associate Director"),
    ("IG Group",                                       "Ram Kumar",                 "Sr. Talent Partner"),
    ("IG Group",                                       "Pooja Singh",               "Global Sr. Talent Partner"),
    ("IG Group",                                       "Shruti Mishra",             "HR Consultant"),
    ("Intuitive AI",                                   "Archi Singhal",             "TA Specialist"),
    ("Intuitive AI",                                   "Guneet Kaur",               "TAG"),
    ("Intuitive AI",                                   "Mehul Vaugh",               "TA Specialist"),
    ("Intuitive AI",                                   "Mitesh Kumar",              "TA Leader"),
    ("Intuitive AI",                                   "Gajalakshmi Ravichandran",  "TA Leader"),
    ("Nucleus Software",                               "Mayank T",                  "TA Director"),
    ("Nucleus Software",                               "Dilip Kumar",               "Global TA"),
    ("Nucleus Software",                               "Chandra Ratra",             "Head of Resource Planning and Talent Acquisition"),
    ("Nucleus Software",                               "Devendra Kashyap",          "TA Manager"),
    ("Nucleus Software",                               "Sachin Sharma",             "Recruitment Specialist"),
    ("Nucleus Software",                               "Deepa Nair",                "Recruitment Lead"),
    ("I Merit",                                        "Tanmoy Saha",               "Lead HR TA"),
    ("I Merit",                                        "Subhanka Chakraborty",      "Sr. Executive TA"),
    ("I Merit",                                        "Sonali Subhadarshini",      "TA Specialist"),
    ("Global Foundries",                               "Divya Chauhan",             "Principal HR Project Manager"),
    ("Global Foundries",                               "Gayatri Manium",            "TA & Business Partner APAC"),
    ("Publicis Sapient",                               "Chetan Nagid",              "Sr. Talent Associate"),
    ("Publicis Sapient",                               "Chetan Sriramwar",          "TA Specialist"),
    ("Publicis Sapient",                               "Sayed Shihabuk Muttaqin",   "Manager TA"),
    ("Publicis Sapient",                               "Shivam Hajare",             "Sr. Associate TA"),
    ("Publicis Sapient",                               "Gourav Soni",               "Sr Associate - TA"),
    ("Publicis Sapient",                               "Sajan Thakur",              "TA Specialist"),
    ("Publicis Sapient",                               "Akhil Saxena",              "Lead TA"),
    ("Coupang",                                        "Divyani Maria",             "Executive Recruitment Coordinator"),
    ("Coupang",                                        "Srinivas Burli",            "Director TA"),
    ("Coupang",                                        "Kanika Jain",               "TA Recruiter"),
    ("Global Data Plc",                                "Rajeev Gupta",              "Director"),
    ("Global Data Plc",                                "Venkata Raman Lagietty",    "Talent Acquisition Leader"),
    ("Global Data Plc",                                "Ramyashree Bomera",         "Sr. Talent Acquisition Specialist"),
    ("Global Data Plc",                                "Tejeswani P",               "Talent Acquisition Partner"),
    ("Razorpay",                                       "Chibhanu Nagri",            "Sr. V.P Peoples Operations"),
    ("Razorpay",                                       "Carol Lucas",               "Associate Director TA"),
    ("Razorpay",                                       "Sumit Premi",               "Sr. Director & Global Head - TA"),
    ("Neilson IQ",                                     "Elanchezhiyan KG",          "Manager TA"),
    ("Neilson IQ",                                     "Simran Mohanty",            "MT - HR"),
    ("Neilson IQ",                                     "Divyashree Dayanand",       "Manager TA"),
    ("Neilson IQ",                                     "Stefen Rosario",            "Associate Director - TA"),
    ("Neilson IQ",                                     "Ayesha Anjum",              "India Lead - TA"),
    ("Guidewire Software",                             "Suparna Das",               "Senior Talent Attraction Partner"),
    ("Guidewire Software",                             "Giriraj Jhala",             "Talent Acquisition Partner"),
    ("Guidewire Software",                             "Anuradha Dash",             "Manager TA"),
    ("Linde",                                          "Jayashree Chandra",         "TA Specialist"),
    ("Linde",                                          "Somdatta Saha",             "TA Specialist"),
    ("Meesho",                                         "Srikarthik Sanaya",         "Head TA"),
    ("Meesho",                                         "Bhasavaraj S Patil",        "Lead Tech Hiring"),
    ("Meesho",                                         "Sachin GS",                 "TA"),
    ("Adobe",                                          "Sanchita HB",               "Talent Architect"),
    ("Morgan Stanley",                                 "Nishat Khan",               "Director TA"),
    ("Ybrant Digital",                                 "Suresh Reddy",              "Founder (Brightcom Group)"),
    ("Trilogy",                                        "Papu Bhatacharya",          "Forward Development Engineer"),
    ("The Hartford",                                   "Manoj Chikala",             "Talent Acquisition Manager"),
    ("Triotree Technologies",                          "Surjeet Thakur",            "Founder COO"),
    ("Apisero Inc",                                    "Navya Krishna Lillari",     "Regional VP - Customer Success & Strategy"),
    ("NTTdata",                                        "Gajendra Menon",            "Sr. Director HR"),
    ("Hexagon",                                        "Navneet Kumar",             "Director"),
    ("Cyient",                                         "Kavita Kurup",              "Chief People Officer"),
    ("Ideagen",                                        "Pranab Kumar Mishra",       "Head of People"),
    ("Fleetx.io",                                      "Anubhav Narula",            "Sr. TA Specialist"),
    ("Vikram Solar",                                   "Kankana Brohmo",            "Associate GM - TA"),
    ("Markets & Markets",                              "Himanni Sharma",            "Asst Manager TA"),
    ("Adani Realty",                                   "Sanjay Chakrabarty",        "Chief Digital Officer"),
    ("LPL Financial",                                  "Naga Rao",                  "Sr. TA Leader"),
    ("IBM",                                            "Sonal Subhadarshini",       "TA Specialist"),
    ("Paychex",                                        "Suhas DS",                  "Director Talent Acquisition"),
    ("Clearwater Analytics",                           "Vivek Verma",               "India Hiring Manager"),
    ("ITC Infotech",                                   "Mansi Adhikari",            "TA Specialist"),
]

def main():
    out_path = Path("data/prospects/input/prospects.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # Header
    ws.append(["S.No", "Client Name", "Poc Name", "Designation", "LinkedIn"])

    # Style header
    from openpyxl.styles import Font, PatternFill, Alignment
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    last_company = ""
    sno          = 1
    for company, person, desig in PROSPECTS:
        display_company = company if company != last_company else ""
        ws.append([sno, display_company, person, desig, ""])
        last_company = company
        sno += 1

    # Column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["E"].width = 40
    ws.freeze_panes = "A2"

    wb.save(str(out_path))
    print(f"Created: {out_path.resolve()}")
    print(f"Total prospects: {len(PROSPECTS)}")

if __name__ == "__main__":
    main()
