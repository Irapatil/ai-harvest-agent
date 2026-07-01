"""
Prospect Excel Export Service — Recruiter Contact Discovery Report

Sheet layout
────────────────────────────────────────────────
Sheet 1 — Lead Intelligence Report   all records, 19 columns
Sheet 2 — Summary                    contact discovery summary

Column spec (19 columns)
────────────────────────────────────────────────────────────────────────
Recruiter Name | Designation | Department | Position Level | Location |
Current Company | LinkedIn Profile URL | Official Email ID | Email Status |
Contact Number | Phone Status | Hiring Domain | Company Industry |
Company Size | Years in Current Company | Overall Experience |
Reporting Manager | Confidence Score | Source
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import structlog

from app.models.prospect_models import ProspectResult

logger = structlog.get_logger(__name__)

_OUTPUT_DIR  = Path("data/results/lead_intelligence")
_ILLEGAL_XML = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f￾￿]')

# ── Column spec: (ProspectResult field name, Excel header label) ──────────────
_COLUMNS: list[tuple[str, str]] = [
    ("person_name",         "Recruiter Name"),
    ("designation",         "Designation"),
    ("department",          "Department"),
    ("position_level",      "Position Level"),
    ("location",            "Location"),
    ("company_name",        "Current Company"),
    ("linkedin_url",        "LinkedIn Profile URL"),
    ("official_email_id",   "Official Email ID"),
    ("email_status",        "Email Status"),
    ("contact_number",      "Contact Number"),
    ("phone_status",        "Phone Status"),
    ("hiring_domain",       "Hiring Domain"),
    ("company_industry",    "Company Industry"),
    ("company_size",        "Company Size"),
    ("years_in_company",    "Years in Current Company"),
    ("overall_experience",  "Overall Experience"),
    ("reporting_manager",   "Reporting Manager"),
    ("confidence_score",    "Confidence Score"),
    ("source",              "Source"),
]

# ── Cell fill colors ──────────────────────────────────────────────────────────

_CONFIDENCE_COLORS = {
    "High":   "C6EFCE",  # green
    "Medium": "FFEB9C",  # amber
    "Low":    "FFC7CE",  # red-pink
}

_EMAIL_STATUS_COLORS = {
    "VERIFIED":  "C6EFCE",  # green  — extracted from corporate directory/website
    "PUBLIC":    "BDD7EE",  # blue   — found on public professional profile
    "NOT_FOUND": "F2F2F2",  # gray
}

_PHONE_STATUS_COLORS = {
    "VERIFIED":  "C6EFCE",  # green
    "PUBLIC":    "BDD7EE",  # blue
    "NOT_FOUND": "F2F2F2",  # gray
}

# ── Summary Sheet rows ────────────────────────────────────────────────────────

def _compute_summary(results: list[ProspectResult]) -> list[tuple[str, int]]:
    return [
        ("Total Recruiters Processed",           len(results)),
        ("Verified Emails Found",                 sum(1 for r in results if r.email_status == "VERIFIED")),
        ("Public Emails Found",                   sum(1 for r in results if r.email_status == "PUBLIC")),
        ("Verified Phones Found",                 sum(1 for r in results if r.phone_status == "VERIFIED")),
        ("Public Phones Found",                   sum(1 for r in results if r.phone_status == "PUBLIC")),
        ("Profiles Without Contact Information",  sum(
            1 for r in results
            if r.email_status == "NOT_FOUND" and r.phone_status == "NOT_FOUND"
        )),
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# Internal helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _sanitize(val: Any) -> Any:
    if isinstance(val, str):
        val = _ILLEGAL_XML.sub("", val)
        if len(val) > 2000:
            val = val[:2000] + "…"
    return val


def _to_row(result: ProspectResult) -> list[Any]:
    return [_sanitize(getattr(result, field, "") or "") for field, _ in _COLUMNS]


def _apply_header_style(ws: Any) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        fill  = PatternFill("solid", fgColor="1F4E79")
        font  = Font(bold=True, color="FFFFFF", size=11)
        align = Alignment(wrap_text=True, vertical="center", horizontal="center")
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align
        ws.row_dimensions[1].height = 30
    except Exception:
        pass


def _color_status_columns(ws: Any) -> None:
    """Color Confidence Score, Email Status, and Phone Status cells by value."""
    try:
        from openpyxl.styles import PatternFill, Font, Alignment

        col_map: dict[str, dict[str, str]] = {}
        for col_idx, (field, _) in enumerate(_COLUMNS, 1):
            if field == "confidence_score":
                col_map[str(col_idx)] = _CONFIDENCE_COLORS
            elif field == "email_status":
                col_map[str(col_idx)] = _EMAIL_STATUS_COLORS
            elif field == "phone_status":
                col_map[str(col_idx)] = _PHONE_STATUS_COLORS

        for row_idx in range(2, ws.max_row + 1):
            for col_idx_str, color_dict in col_map.items():
                cell  = ws.cell(row=row_idx, column=int(col_idx_str))
                value = str(cell.value or "").strip()
                color = color_dict.get(value)
                if color:
                    cell.fill      = PatternFill("solid", fgColor=color)
                    cell.font      = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
    except Exception:
        pass


def _autofit(ws: Any, max_width: int = 55) -> None:
    try:
        from openpyxl.utils import get_column_letter
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in col_cells), default=12)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, max_width)
    except Exception:
        pass


def _write_data_sheet(ws: Any, results: list[ProspectResult]) -> None:
    ws.append([label for _, label in _COLUMNS])
    for r in results:
        ws.append(_to_row(r))
    _apply_header_style(ws)
    _color_status_columns(ws)
    _autofit(ws)
    ws.freeze_panes = "A2"


def _write_summary_sheet(ws: Any, results: list[ProspectResult]) -> None:
    try:
        from openpyxl.styles import Font, PatternFill, Alignment

        ws.append(["Metric", "Count"])

        for metric, count in _compute_summary(results):
            ws.append([metric, count])

        # Header style
        fill  = PatternFill("solid", fgColor="1F4E79")
        font  = Font(bold=True, color="FFFFFF", size=11)
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = align
        ws.row_dimensions[1].height = 30

        # Count column — right-aligned, bold
        for row_idx in range(2, ws.max_row + 1):
            cell           = ws.cell(row=row_idx, column=2)
            cell.font      = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        _autofit(ws, max_width=60)
        ws.freeze_panes = "A2"
    except Exception as exc:
        logger.warning("summary_sheet_style_failed", error=str(exc))


# ═══════════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════════

class ProspectExcelService:
    """Generate Lead_Intelligence_Report.xlsx from enriched prospect results."""

    def export(
        self,
        results: list[ProspectResult],
        run_id: str,
        report_title: str = "Lead Intelligence Report",
    ) -> str:
        try:
            import openpyxl
        except ImportError:
            logger.error("prospect_excel_skipped", reason="openpyxl not installed")
            return ""

        _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        filename = f"{run_id}_Lead_Intelligence_Report.xlsx"
        path     = _OUTPUT_DIR / filename

        wb = openpyxl.Workbook()

        # Sheet 1 — All records (consolidated)
        ws_all       = wb.active
        ws_all.title = report_title[:31]   # Excel max sheet name is 31 chars
        _write_data_sheet(ws_all, results)

        # Sheet 2 — Summary
        ws_summary = wb.create_sheet(title="Summary")
        _write_summary_sheet(ws_summary, results)

        wb.save(str(path))

        logger.info(
            "prospect_excel_exported",
            path            = str(path.resolve()),
            total           = len(results),
            verified_emails = sum(1 for r in results if r.email_status == "VERIFIED"),
            public_emails   = sum(1 for r in results if r.email_status == "PUBLIC"),
            verified_phones = sum(1 for r in results if r.phone_status == "VERIFIED"),
            public_phones   = sum(1 for r in results if r.phone_status == "PUBLIC"),
            high            = sum(1 for r in results if r.confidence_score == "High"),
            medium          = sum(1 for r in results if r.confidence_score == "Medium"),
            low             = sum(1 for r in results if r.confidence_score == "Low"),
        )
        return str(path.resolve())
